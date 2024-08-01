import copy

from openpyxl import load_workbook


titiles = ["定位历元数对比", "定位历元数占比对比", "水平各项指标对比表", "垂向各项指标对比表"]
titiles_idx = [None, None, None, None]


def set_titiles_idx(idx):
    global titiles_idx

    tmp_idx = 0
    while True:
        if tmp_idx == 4:
            break
        if titiles_idx[tmp_idx] is None:
            titiles_idx[tmp_idx] = idx
            break
        tmp_idx += 1

def get_ref_title_name():
    global titiles
    return titiles.pop(0)


def check_excel_format(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    print(f"max row: {sheet.max_row}    max column: {sheet.max_column}")

    rows = sheet.iter_rows(min_row=1, max_row=sheet.max_row)
    ref_title_name = get_ref_title_name()
    for row in rows:
        title_name = row[0].value
        if not title_name == ref_title_name:
            if row[0].row == sheet.max_row:
                raise Exception(f"{row.row}行数据格式错误, 目标内容为: {ref_title_name}")
            else:continue
        else:
            set_titiles_idx(copy.deepcopy(row[0].row))

            try:
                ref_title_name = get_ref_title_name()
            except IndexError:
                break

    print(titiles_idx)

def main():
    xlsxFilePath = r"C:\Users\12587\Desktop\BK\Comparison_2_BK_1662_L1L5_23102.xlsx"
    check_excel_format(xlsxFilePath)


if __name__ == '__main__':
    main()