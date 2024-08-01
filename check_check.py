# -*- encoding: utf-8 -*-
# 注意需要剔除多余的文件夹，搜索notice
import copy
import os
import Fuction
from loguru import logger
from interval import Interval
from Fuction import utc_to_sec
import pynmea2


def check_rtk_delay():
    incomplete_data = list()
    try:
        f_path = input('raw/nmea file path: ')
        output_data_fpath = os.path.dirname(f_path) + '\\results.txt'
        incomplete_data_path = os.path.dirname(f_path) + '\\incomplete_data.txt'
        raw_f = open(f_path, 'r', encoding='utf-8', errors='ignore')
        output_f = open(output_data_fpath, 'w+', encoding='utf-8')
        # incomplete_dataf = open(incomplete_data_path, 'w+', encoding='utf-8')
        for line in raw_f.readlines():
            if line.startswith('$GNGGA'):
                cnt_comma = line.count(',')
                if cnt_comma != 14:
                    incomplete_data.append(line)
                    print(f'GGA不完整： {line}')
                else:
                    time, delay = line.split(',')[1], line.split(',')[-2]
                    output_f.write(f"{time},{delay}\n")

        raw_f.close()
        output_f.close()
        # for line in incomplete_data:
        #     incomplete_dataf.write(f"{line}\n")
        print('end')
    except Exception as e:
        logger.exception(e)


def check_tunnel_length():
    cases_dic = dict()
    try:
        # source_root = input('info file root: ')
        source_root = r'C:\mengfo.wang\test\data\SPPReplayAnalyze\test_kml'
        out_data_fpath = source_root + '\\results.txt'

        out_f = open(out_data_fpath, 'a', encoding='utf-8')
        # 循环case
        for root, dirs, files in os.walk(source_root):
            for casefile in dirs:
                # notice
                if 'FILE' not in casefile:  # 剔除多余的文件夹
                    continue
                cases_dic[casefile] = dict()

                cur_case_root = os.path.join(root, casefile)

                fileNfolder_names = os.listdir(cur_case_root)

                # 剔除文件夹，只保留当前文件夹下的文件名notice
                while True:
                    check_change = copy.deepcopy(fileNfolder_names)
                    for name in fileNfolder_names:
                        if '.' not in name:
                            fileNfolder_names.remove(name)
                            break
                        if 'info' not in name:
                            fileNfolder_names.remove(name)
                            break
                    if len(check_change) == len(fileNfolder_names):
                        break
                cur_files = fileNfolder_names

                # 循环info文件
                for chip_name in cur_files:

                    cur_info_f = os.path.join(cur_case_root, chip_name)

                    cases_dic[casefile][chip_name] = list()

                    tunnels: list = cases_dic[casefile][chip_name]

                    cnt_tunnel = 0

                    info_fopen = open(cur_info_f, 'r', encoding='utf-8')

                    for line in info_fopen.readlines():
                        if '$Tunnel: ' in line:
                            cnt_tunnel += 1

                            new_line = line.replace('$Tunnel: ', '')
                            data_list = new_line.split(',')
                            start_time, end_time, length = data_list[0:3]

                            tunnels.append([cnt_tunnel, start_time, end_time, length])

        for casefile in cases_dic:
            for chip_name in cases_dic[casefile]:
                bk_cnt_tn = 0
                other_cnt_tn = 0
                if 'BK' in chip_name or 'bK' in chip_name:
                    bk_cnt_tn = len(cases_dic[casefile][chip_name])
                    for other_chip_name in cases_dic[casefile]:
                        if 'BK' not in other_chip_name and 'K801' in other_chip_name:
                            other_cnt_tn = len(cases_dic[casefile][other_chip_name])
                            if bk_cnt_tn >= other_cnt_tn:
                                continue
                            else:
                                out_f.write(f'tunnel数量错误{casefile} {chip_name}, {other_chip_name} {bk_cnt_tn} {other_cnt_tn}\n')
                                print(f'tunnel数量错误{casefile} {chip_name}, {other_chip_name} {bk_cnt_tn} {other_cnt_tn}')
                                break

        for casefile in cases_dic:
            for chip_name in cases_dic[casefile]:
                if 'BK' in chip_name or 'bK' in chip_name:
                    for tunnel in cases_dic[casefile][chip_name]:
                        if tunnel:
                            bk_tn_id, bk_tn_start, bk_tn_end, bk_tn_length = tunnel[0], tunnel[1], tunnel[2], tunnel[3]
                            for other_chip_name in cases_dic[casefile]:
                                if 'BK' not in other_chip_name and 'K801' in other_chip_name:
                                    for other_tunnel in cases_dic[casefile][other_chip_name]:
                                        if other_tunnel:
                                            other_tn_start, other_tn_end, other_tn_length = \
                                                other_tunnel[1], other_tunnel[2], other_tunnel[3]
                                            if utc_to_sec(other_tn_start) - utc_to_sec(bk_tn_start) in Interval(-10, 10):  # 对齐tunnel
                                                # check 1
                                                if float(bk_tn_length) - float(other_tn_length) > 5:
                                                    print(f'tunnle时间太长：{casefile} {chip_name} {bk_tn_id},'
                                                          f'{float(bk_tn_length) - float(other_tn_length)}')
                                                # check 2
                                                if utc_to_sec(bk_tn_end) - utc_to_sec(other_tn_end) > 4:
                                                    out_f.write(f'出tunnle定位慢：{casefile}\t{chip_name}\t{bk_tn_id},'
                                                                f'\t{utc_to_sec(bk_tn_end) - utc_to_sec(other_tn_end)}\n')
                                                    print(f'出tunnle定位慢：{casefile}\t{chip_name}\t{bk_tn_id},'
                                                          f'\t{utc_to_sec(bk_tn_end) - utc_to_sec(other_tn_end)}')
                                                # else:
                                                #     print(f'出tunnle定位正常：{casefile}\t{chip_name}\t{bk_tn_id},'
                                                #           f'\t{utc_to_sec(bk_tn_end) - utc_to_sec(other_tn_end)}\n')
                                        else: print(f'{casefile}\t{chip_name}没有tunnel！')
        print('end check_tunnel.\n')
    except Exception as e:
        logger.exception(e)


def check_basement_length():
    cases_dic = dict()
    try:
        # source_root = input('info file root: ')
        source_root = r'C:\mengfo.wang\test\data\SPPReplayAnalyze\test_kml'
        out_data_fpath = source_root + '\\results.txt'

        out_f = open(out_data_fpath, 'a', encoding='utf-8')
        # 循环case
        for root, dirs, files in os.walk(source_root):
            for casefile in dirs:
                # notice
                if 'FILE' not in casefile:  # 剔除多余的文件夹
                    continue
                cases_dic[casefile] = dict()
                cur_case_root = os.path.join(root, casefile)
                fileNfolder_names = os.listdir(cur_case_root)

                # 剔除文件夹，只保留当前文件夹下的文件名*****notice*****
                while True:
                    check_change = copy.deepcopy(fileNfolder_names)
                    for name in fileNfolder_names:
                        if '.' not in name:
                            fileNfolder_names.remove(name)
                            break
                        if 'info' not in name:
                            fileNfolder_names.remove(name)
                            break
                    if len(check_change) == len(fileNfolder_names):
                        break
                cur_files = fileNfolder_names

                # 循环info文件
                for chip_name in cur_files:

                    cur_info_f = os.path.join(cur_case_root, chip_name)

                    basements = list()

                    cases_dic[casefile][chip_name] = dict()

                    cases_dic[casefile][chip_name]['basements'] = basements

                    cnt_basement = 0

                    info_fopen = open(cur_info_f, 'r', encoding='utf-8')

                    for line in info_fopen.readlines():
                        if '地库' in line:
                            cnt_basement += 1

                            new_line = line.replace('$Outage: ', '')
                            data_list = new_line.split(',')
                            start_time, end_time, length = data_list[0:3]

                            basements.append([cnt_basement, start_time, end_time, length])

        for casefile in cases_dic:
            for chip_name in cases_dic[casefile]:
                bk_cnt_tn = 0
                other_cnt_tn = 0
                if 'BK' in chip_name or 'bK' in chip_name:
                    bk_cnt_tn = len(cases_dic[casefile][chip_name]['basements'])
                    for other_chip_name in cases_dic[casefile]:
                        if 'BK' not in other_chip_name and 'K801' in other_chip_name:  # 选取参考对象
                            other_cnt_tn = len(cases_dic[casefile][other_chip_name]['basements'])
                            if bk_cnt_tn >= other_cnt_tn:
                                continue
                            else:
                                out_f.write(f'地库数量错误{casefile} {chip_name}, {other_chip_name} {bk_cnt_tn} {other_cnt_tn}\n')
                                print(f'地库数量错误{casefile} {chip_name}, {other_chip_name} {bk_cnt_tn} {other_cnt_tn}')
                                break

        for casefile in cases_dic:
            for chip_name in cases_dic[casefile]:
                if 'BK' in chip_name or 'bK' in chip_name:
                    for basement in cases_dic[casefile][chip_name]['basements']:
                        if basement:
                            bk_bs_id, bk_bs_start, bk_bs_end, bk_bs_length = basement[0], basement[1], basement[2], basement[3]
                            for other_chip_name in cases_dic[casefile]:
                                if 'BK' not in other_chip_name and 'K801' in other_chip_name:  # 选取参考对象
                                    for other_bs in cases_dic[casefile][other_chip_name]['basements']:
                                        if other_bs:
                                            other_bs_start, other_bs_end, other_bs_length = \
                                                other_bs[1], other_bs[2], other_bs[3]
                                            if utc_to_sec(other_bs_start) - utc_to_sec(bk_bs_start) in Interval(-10, 10):  # 对齐tunnel
                                                # check 1
                                                # if float(bk_tn_length) - float(other_tn_length) > 4:
                                                #     print(f'tunnle时间太长：{case} {chip_name} {bk_tn_id},'
                                                #           f'{float(bk_tn_length) - float(other_tn_length)}')
                                                # check 2
                                                if utc_to_sec(bk_bs_end) - utc_to_sec(other_bs_end) > 4:
                                                    out_f.write(f'出地库定位慢：{casefile}\t{chip_name}\t{bk_bs_end},'
                                                                f'\t{utc_to_sec(bk_bs_end) - utc_to_sec(other_bs_end)}\n')
                                                    print(f'出地库定位慢：{casefile}\t{chip_name}\t{bk_bs_end},'
                                                          f'\t{utc_to_sec(bk_bs_end) - utc_to_sec(other_bs_end)}')
        print('end check_basement.\n')
    except Exception as e:
        logger.exception(e)


def check_tnNbs_err(print_all_tn=False, print_all_bs=False):
    """
    当print_all_tn=True时不处理tunnel数据，只打印tunnel数据；print_all_bs同理。
    :param print_all_tn: 扫描所有场景，打印tunnel数据
    :param print_all_bs: 扫描所有场景，打印basement数据
    :return:
    """
    cases_dic = dict()
    try:
        # source_root = input('info file root: ')
        source_root = r'C:\mengfo.wang\test\data\SPPReplayAnalyze\test_kml'
        out_data_fpath = source_root + '\\results.txt'

        out_f = open(out_data_fpath, 'a', encoding='utf-8')
        # 循环case
        for root, dirs, files in os.walk(source_root):
            for casefile in dirs:
                # notice
                cases_dic[casefile] = dict()

                cur_case_root = os.path.join(root, casefile)

                if not os.path.isdir(cur_case_root):  # 剔除多余的文件夹
                    continue

                fileNfolder_names = os.listdir(cur_case_root)

                # 剔除文件夹，只保留当前文件夹下的文件名notice
                while True:
                    check_change = copy.deepcopy(fileNfolder_names)
                    for name in fileNfolder_names:
                        if '.' not in name:
                            fileNfolder_names.remove(name)
                            break
                        if 'info' not in name:
                            fileNfolder_names.remove(name)
                            break
                    if len(check_change) == len(fileNfolder_names):
                        break
                cur_files = fileNfolder_names

                # 循环info文件
                for chip_name in cur_files:

                    cur_info_f = os.path.join(cur_case_root, chip_name)

                    # 初始化
                    cases_dic[casefile][chip_name] = dict()
                    cases_dic[casefile][chip_name]['tunnels'] = list()
                    tunnels = cases_dic[casefile][chip_name]['tunnels']
                    cnt_tunnel = 0
                    cases_dic[casefile][chip_name]['basements'] = list()
                    basements = cases_dic[casefile][chip_name]['basements']
                    cnt_basement = 0

                    info_fopen = open(cur_info_f, 'r', encoding='utf-8')
                    for line in info_fopen.readlines():
                        if '$Tunnel: ' in line:
                            cnt_tunnel += 1

                            new_line = line.replace('$Tunnel: ', '')

                            data_list = new_line.split(',')

                            start_time, end_time, length = data_list[0:3]

                            tunnels.append([cnt_tunnel, start_time, end_time, length])

                        if '地库' in line:
                            cnt_basement += 1

                            new_line = line.replace('$Outage: ', '')

                            data_list = new_line.split(',')

                            start_time, end_time, length = data_list[0:3]

                            basements.append([cnt_basement, start_time, end_time, length])

        for casefile in cases_dic:

            for chip_name in cases_dic[casefile]:

                if 'BK' in chip_name or 'bK' in chip_name:

                    for tunnel in cases_dic[casefile][chip_name]['tunnels']:

                        if tunnel:

                            bk_tn_id, bk_tn_start, bk_tn_end, bk_tn_length = tunnel[0], tunnel[1], tunnel[2], tunnel[3]

                            if print_all_tn:
                                print(f"{casefile}\t{chip_name}\t{bk_tn_id},{bk_tn_start},{bk_tn_end},{bk_tn_length}")

                            else:
                                tn_file = os.path.join(source_root, casefile, chip_name)

                                f_open = open(tn_file, mode='r', encoding='utf-8')

                                for line in f_open.readlines():

                                    if not line.startswith('$'):

                                        line = line.rstrip('\n')

                                        time, err = line.split(',')[0], float(line.split(',')[1])

                                        # if time == bk_tn_end:
                                        if utc_to_sec(time) - utc_to_sec(bk_tn_end) in Interval(0, 10):  # 匹配出tunnel10秒内的数据
                                            if err not in Interval(-50, 50):
                                                print(f'再次定位误差过大(隧道)! {casefile} {chip_name} time/err: {time} {err}')

                                            if utc_to_sec(time) - utc_to_sec(bk_tn_end) == 10:  # 跳过出tunnel10秒后的数据
                                                break
                                f_open.close()

                    for basement in cases_dic[casefile][chip_name]['basements']:

                        if basement:

                            bk_bs_id, bk_bs_start, bk_bs_end, bk_bs_length = basement[0], basement[1], basement[2], basement[3]

                            if print_all_bs:
                                print(f"{casefile}\t{chip_name}\t{bk_bs_id},{bk_bs_start},{bk_bs_end},{bk_bs_length}")
                            else:
                                bs_file = os.path.join(source_root, casefile, chip_name)

                                f_open = open(bs_file, mode='r', encoding='utf-8')

                                for line in f_open.readlines():

                                    if not line.startswith('$'):

                                        line = line.rstrip('\n')

                                        time, err = line.split(',')[0], float(line.split(',')[1])

                                        if utc_to_sec(time) - utc_to_sec(bk_bs_end) in Interval(0, 10):

                                            if err not in Interval(-50, 50):
                                                print(f'再次定位误差过大（地库）! {casefile} {chip_name} time/err: {time} {err}\n')

                                            if utc_to_sec(time) - utc_to_sec(bk_bs_end) == 10:  # 跳过出tunnel10秒后的数据
                                                break
                                f_open.close()
        print('end check_tnNbs_err.\n')
    except Exception as e:
        logger.exception(e)


def check_TTFFandERR():
    print("check ttff here")


def delete_exgga():
    """
    剔除GIGGA
    :return:
    """
    try:
        f_path = input('raw/nmea file path: ')
        output_data_fpath = os.path.dirname(f_path) + '\\results.txt'
        raw_f = open(f_path)
        output_f = open(output_data_fpath, 'w+', encoding='utf-8')
        for line in raw_f.readlines():
            if line.startswith('$GIGGA'):
                continue
            else:
                output_f.write(line)
        raw_f.close()
        output_f.close()
        print('end')
    except Exception as e:
        logger.exception(e)


def cnt_cn0(f_path, process_time, case, file):
    try:
        # f_path = input('raw/nmea file path: ')
        # f_path = r"D:\mfw\tmp\5_BK_DBG_L1_Cold_None_fw_info.log"
        # process_time = input('time:')
        # process_time = '223004'
        output_data_fpath = os.path.dirname(f_path) + '\\results.txt'
        raw_f = open(f_path, mode='r', encoding='utf-8', errors='ignore')
        output_f = open(output_data_fpath, 'a+', encoding='utf-8')
        output_f.write(f_path + '\n')

        opo_list = list()
        sys_cn0 = list()
        gsv_marker = ''
        into_gsv = False
        # try:
        #     for line in raw_f.readlines():
        #         a = line
        # except:
        #     raw_f.close()
        #     raw_f = open(f_path, mode='r', encoding='utf-8', errors='ignore')
        #     pass
        for line in raw_f.readlines():
            if line.startswith('$') or '$GNGGA' in line:
                if line.startswith('$GNGGA') or '$GNGGA' in line:
                    line = line.split('$GNGGA')[1]
                    line = '$GNGGA' + line
                    if into_gsv:
                        if cur_cnt_cn0 != 0 and cur_cnt_sat != 0:
                            # for satinfo in gsv_list:
                            #     for sys, satnum, satcn0 in satinfo:
                            #         if sys == gsv_marker:
                            #             now_check =

                            cur_avg_cn0 = cur_cnt_cn0 / cur_cnt_sat
                            sys_cn0.append([gsv_marker, cur_cnt_sat, cur_avg_cn0])
                        break
                    if process_time not in line:
                        continue
                    else:
                        into_gsv = True
                    if line.count(',') != 14:
                        print(f'gga数据不完整！')
                        # print(f'gga数据不完整！ {line}')
                    cur_time = line.split(',')[1]
                    cur_gsa_sat = line.split(',')[7]

                    if process_time in cur_time:
                        gsv_list = list()
                        gsv_marker = ''
                        pass
                    else:continue
                #############################################
                if into_gsv:
                    if line.startswith('$GPGSV'):
                        if gsv_marker == '':
                            cur_cnt_cn0 = 0
                            cur_cnt_sat = 0
                            gsv_marker = '$GPGSV'
                        else:
                            if gsv_marker == '$GPGSV':
                                pass
                            else:
                                if cur_cnt_cn0 != 0 and cur_cnt_sat != 0:
                                    cur_avg_cn0 = cur_cnt_cn0 / cur_cnt_sat
                                    sys_cn0.append([gsv_marker, cur_cnt_sat, cur_avg_cn0])
                                cur_cnt_cn0 = 0
                                cur_cnt_sat = 0
                                gsv_marker = '$GPGSV'
                        if line.count(',') != 20:
                            print(f'gsv数据不完整！ {line}')

                        if line.count(',') < 8:
                            continue
                        cur_sat = line.split(',')[4]
                        cur_cn0 = line.split(',')[7]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GPGSV', cur_sat, cur_cn0])
                        if line.count(',') < 11:
                            continue
                        cur_sat = line.split(',')[8]
                        cur_cn0 = line.split(',')[11]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GPGSV', cur_sat, cur_cn0])
                        if line.count(',') < 15:
                            continue
                        cur_sat = line.split(',')[12]
                        cur_cn0 = line.split(',')[15]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GPGSV', cur_sat, cur_cn0])
                        if line.count(',') < 19:
                            continue
                        cur_sat = line.split(',')[16]
                        cur_cn0 = line.split(',')[19]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GPGSV', cur_sat, cur_cn0])
                #############################################
                if into_gsv:
                    if line.startswith('$BDGSV') or line.startswith('$GBGSV'):
                        # 只统计L1信号
                        end_of_this_line = line.split(',')[-1]
                        sig_id = end_of_this_line.split('*')[0]
                        if sig_id != '1':
                            print(f'signal ID 不为 1：{line}')
                            continue
                        else:
                            pass
                        if gsv_marker == '':
                            cur_cnt_cn0 = 0
                            cur_cnt_sat = 0
                            gsv_marker = '$BDGSV'
                        else:
                            if gsv_marker == '$BDGSV':
                                pass
                            else:
                                if cur_cnt_cn0 != 0 and cur_cnt_sat != 0:
                                    cur_avg_cn0 = cur_cnt_cn0 / cur_cnt_sat
                                    sys_cn0.append([gsv_marker, cur_cnt_sat, cur_avg_cn0])
                                cur_cnt_cn0 = 0
                                cur_cnt_sat = 0
                                gsv_marker = '$BDGSV'
                        if line.count(',') != 20:
                            print(f'gsv数据不完整！ {line}')

                        if line.count(',') < 8:
                            continue
                        cur_sat = line.split(',')[4]
                        cur_cn0 = line.split(',')[7]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['BDGSV', cur_sat, cur_cn0])
                        if line.count(',') < 11:
                            continue
                        cur_sat = line.split(',')[8]
                        cur_cn0 = line.split(',')[11]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['BDGSV', cur_sat, cur_cn0])
                        if line.count(',') < 15:
                            continue
                        cur_sat = line.split(',')[12]
                        cur_cn0 = line.split(',')[15]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['BDGSV', cur_sat, cur_cn0])
                        if line.count(',') < 19:
                            continue
                        cur_sat = line.split(',')[16]
                        cur_cn0 = line.split(',')[19]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['BDGSV', cur_sat, cur_cn0])
                #############################################
                if into_gsv:
                    if line.startswith('$GLGSV'):
                        if gsv_marker == '':
                            cur_cnt_cn0 = 0
                            cur_cnt_sat = 0
                            gsv_marker = '$GLGSV'
                        else:
                            if gsv_marker == '$GLGSV':
                                pass
                            else:
                                if cur_cnt_cn0 != 0 and cur_cnt_sat != 0:
                                    cur_avg_cn0 = cur_cnt_cn0 / cur_cnt_sat
                                    sys_cn0.append([gsv_marker, cur_cnt_sat, cur_avg_cn0])
                                cur_cnt_cn0 = 0
                                cur_cnt_sat = 0
                                gsv_marker = '$GLGSV'
                        if line.count(',') != 20:
                            print(f'gsv数据不完整！ {line}')

                        if line.count(',') < 8:
                            continue
                        cur_sat = line.split(',')[4]
                        cur_cn0 = line.split(',')[7]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GLGSV', cur_sat, cur_cn0])
                        if line.count(',') < 11:
                            continue
                        cur_sat = line.split(',')[8]
                        cur_cn0 = line.split(',')[11]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GLGSV', cur_sat, cur_cn0])
                        if line.count(',') < 15:
                            continue
                        cur_sat = line.split(',')[12]
                        cur_cn0 = line.split(',')[15]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GLGSV', cur_sat, cur_cn0])
                        if line.count(',') < 19:
                            continue
                        cur_sat = line.split(',')[16]
                        cur_cn0 = line.split(',')[19]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GLGSV', cur_sat, cur_cn0])
                #############################################
                if into_gsv:
                    if line.startswith('$GAGSV'):
                        if gsv_marker == '':
                            cur_cnt_cn0 = 0
                            cur_cnt_sat = 0
                            gsv_marker = '$GAGSV'
                        else:
                            if gsv_marker == '$GAGSV':
                                pass
                            else:
                                if cur_cnt_cn0 != 0 and cur_cnt_sat != 0:
                                    cur_avg_cn0 = cur_cnt_cn0/cur_cnt_sat
                                    sys_cn0.append([gsv_marker, cur_cnt_sat, cur_avg_cn0])
                                cur_cnt_cn0 = 0
                                cur_cnt_sat = 0
                                gsv_marker = '$GAGSV'
                        if line.count(',') != 20:
                            print(f'gsv数据不完整！ {line}')

                        if line.count(',') < 8:
                            continue
                        cur_sat = line.split(',')[4]
                        cur_cn0 = line.split(',')[7]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GAGSV', cur_sat, cur_cn0])
                        if line.count(',') < 11:
                            continue
                        cur_sat = line.split(',')[8]
                        cur_cn0 = line.split(',')[11]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GAGSV', cur_sat, cur_cn0])
                        if line.count(',') < 15:
                            continue
                        cur_sat = line.split(',')[12]
                        cur_cn0 = line.split(',')[15]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GAGSV', cur_sat, cur_cn0])
                        if line.count(',') < 19:
                            continue
                        cur_sat = line.split(',')[16]
                        cur_cn0 = line.split(',')[19]
                        if cur_cn0:
                            if float(cur_cn0):
                                cur_cnt_cn0 += float(cur_cn0)
                                cur_cnt_sat += 1
                                gsv_list.append(['GAGSV', cur_sat, cur_cn0])
                else:
                    continue
            else:
                pass
        for sys in sys_cn0:
            for data in sys:
                output_f.write(str(data))
                if sys.index(data)+1 == len(sys):
                    output_f.write('\n')
                else:
                    output_f.write(',')
        total_cn0 = 0
        for sat in gsv_list:
            total_cn0 += float(sat[2])
        avg_toal_cn0 = total_cn0 / len(gsv_list)

        output_f.write(f'{str(len(gsv_list))}, {avg_toal_cn0}\n')
        raw_f.close()
        output_f.close()
        print('end')
    except Exception as e:
        logger.exception(e)


def get_cn0():
    # files = [
    #     '1_BK_DBG_L1L5_Hot_0010_fw_info.log', '2_BK_DBG_L1_COLD_0011_fw_info.log', '3_UM620N_L1L5_Hot.dat',
    #     '4_SKG123L_L1L5_Hot.dat', '5_BK_DBG_L1_Cold_None.dat', '6_3352_L1_Cold.dat', '7_UM220_L1_Cold.dat'
    files = [
        '3352_L1_COLD.dat', 'BK_23A_L1_COLD_fw_info.log', '3352_L1_Cold.dat'
    ]
    case = 'BK_FILE033_HalfSky_PowerOff_10'
    time = '122045'

    # 实网
    # time1 = '072626'
    # time2 = '081220'
    # time3 = '104154'
    # time4 = '122045'  file033强信号

    # 模拟器
    # time1 = '221157'
    # time2 = '221338'
    # time3 = '221238'
    # time4 = '221317'

    target_root = r'D:\mfw\simtest\halfsky'
    # for case in os.listdir(target_root):
    #     case_root = os.path.join(target_root, case)
    #     for file in os.listdir(case_root):
    #         if file in files:
    #             file_path = os.path.join(case_root, file)
    #             cnt_cn0(file_path, time, case, file)
    case_root = os.path.join(target_root, case)
    for file in os.listdir(case_root):
        if file in files:
            file_path = os.path.join(case_root, file)
            cnt_cn0(file_path, time, case, file)


def do_xl():
    import openpyxl

    xlfile = r"C:\mengfo.wang\test\data\SPPReplayAnalyze\test_data_src\case_FILE038_Dynamic\Compared\case_FILE038_Dynamic.xlsx"

    # 打开 Excel 文件,不存在则报错
    workbook = openpyxl.load_workbook(xlfile)

    # 获取当前活动工作表（默认为第一张）
    sheet = workbook.active

    # 遍历每一行并输出 A 列的值
    for cur_row in range(4, 12):
        for row in sheet.iter_rows(min_row=cur_row, max_row=cur_row, min_col=2, max_col=sheet.max_column, values_only=True):
            for cell in row:
                print(cell)

    # 创建新的 Excel 文件并添加数据
    new_workbook = openpyxl.Workbook()  # 创建新的excel
    new_sheet = new_workbook.active      # 获取当前活动表
    new_sheet["A1"] = "Hello, World!"      # 在表格 A1 位置写入数据 "Hello, World!"
    new_workbook.save(xlfile)            # 保存表格


def check_epo(xlsxfile):
    import openpyxl

    # 存储xlsx数据的字典
    data = dict()

    # 打开目标xlsx文件
    workbook = openpyxl.load_workbook(xlsxfile)

    # 获取当前活动sheet
    sheet = workbook.active

    # 获取历元数据dict
    epo_data = get_epo_num(sheet)

    #
    for chipname in epo_data:
        if 'K801' in chipname:
            k801_eponum = epo_data[chipname]
            break

    for chipname in epo_data:
        if 'BK' in chipname:
            bk_eponum = epo_data[chipname]

        if bk_eponum and k801_eponum:
            if bk_eponum - k801_eponum >= 50:
                print(f"历元数偏少：{bk_eponum - k801_eponum}")
        else:
            raise Exception("数据不存在")


def get_epo_num(sheet):
    # 存储xlsx数据的字典
    data = dict()

    # 行遍历
    for col in sheet.iter_cols(min_row=3, min_col=2, max_row=10, max_col=3, values_only=True):

        # 遇到空白行，重置keyflag = True
        if len(col) == col.count(None):
            continue

        # 初始化data[chipname]字典
        if not data:
            for cell in col:
                if cell is None:
                    continue
                else:
                    data[cell] = list()
                    continue

        # 填充data[chipname]字典
        else:
            for cell in col:
                if cell is None:
                    continue
                else:
                    data = dict(sorted(data.items()))
                    for item in data:
                        if data[item]:
                            continue
                        else:
                            data[item] = cell
                            continue
                    continue

    return data


def get_tag(sheet):
    """
    行遍历表格获取表格的key
    :param sheet:
    :return:
    """
    # 存储xlsx数据的字典
    data = dict()

    keyflag = True  # 遇到空白行会重置keyflag为True，遇到非空白行的第一行之后置False，以此达到将表格属性名（key）和属性值进行区分

    chartID = 0  # 不同的表格用不同的chartID标记

    # 行遍历
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True):

        # 遇到空白行，重置keyflag = True
        if len(row) == row.count(None):

            keyflag = True

            continue

        # 初始化data[key]字典
        if keyflag:

            for cell in row:

                if cell is None:

                    continue

                else:
                    # 规避名称重复的属性名（key）
                    if cell in data:
                        continue
                        # raise Exception(f'重复的key:{cell}')

                    data[cell] = dict()

            keyflag = False

        else:
            continue
    return data


def get_value(sheet):
    """
    列遍历表格获取表格的value
    :param sheet:
    :return:
    """
    # 存储xlsx数据的字典
    data = dict()

    valueflag = False

    chartID = 0  # 不同的表格用不同的chartID标记

    # 行遍历
    for col in sheet.iter_cols(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True):

        if len(col) == col.count(None):
            continue

        # 初始化data[key]字典
        for cell in col:

            if cell is None:

                continue

            else:
                if cell in data:
                    valueflag = True
                    continue
                    # raise Exception(f'重复的key:{cell}')

                data[cell] = dict()


        else:
            continue
    return data


def spp_check():
    check_basement_length()

    check_tunnel_length()

    check_tnNbs_err()


def change_nmea_value(content, talkerID, pos):
    """

    :param talkerID:
    :param pos:
    :return:
    """
    GNGGA_cnt_comma = 14
    GNGSA_cnt_comma = 18
    GNRMC_cnt_comma = 13

    rawfile_dir = r'D:\mfw\1_test\111'
    for iterm in os.listdir(rawfile_dir):
        if '.nmea' in iterm:

            # 源文件
            srcfile = os.path.join(rawfile_dir, iterm)
            src = open(srcfile, 'r', encoding='utf-8', errors='ignore')

            # 结果文件
            rs_newiterm = 'rs_' + iterm  # 为了在名称上区分结果文件和源文件
            rsfile = os.path.join(rawfile_dir, rs_newiterm)
            rs = open(rsfile, 'a', encoding='utf-8', errors='ignore')

            # 遍历行
            for line in src.readlines():
                line = line.rstrip('\n')  # 去掉句末的换行符号

                # 存储替换后生成的新nmea语句
                newline = ''

                # 若该行为目标nmea语句
                if talkerID in line:

                    # 替换对应的nmea语句
                    if talkerID == 'GNGGA':
                        cnt_comma = line.count(',')
                        if cnt_comma != GNGGA_cnt_comma:
                            print(f"数据缺失！\n{line}")
                        else:
                            infolist = line.split(',')
                            infolist[pos] = content
                            newline = ','.join(infolist)

                    if talkerID == 'GNRMC':
                        cnt_comma = line.count(',')
                        if cnt_comma != GNRMC_cnt_comma:
                            print(f"数据缺失！\n{line}")
                        else:
                            infolist = line.split(',')
                            infolist[pos] = content
                            newline = ','.join(infolist)

                if not newline:
                    newline = line

                rs.write(newline)

            src.close()
            rs.close()


def search(wordmark, filemark):
    '''
    在filemark路径下，批量搜索关键字wordmark，打印搜索过的文件，打印wordmark所在行。
    :param wordmark: 搜索关键字
    :param filemark: 被搜索的文件名特征
    :return:
    '''

    from colorama import Fore, Back, Style

    try:
        source_root = input('info file root: ')
        # source_root = r'\\192.168.60.72\Test\TestData(72)\166023295\data_spp_replay'
        # source_root = r'\\192.168.60.217\Exchang_in\Common\WTR_LOG\23.07.03_166023265\稳定性板子重启'
        # source_root = r'D:\mfw\1_test\tracking_nmea\2023717'

        for root, dirs, files in os.walk(source_root):
            for file in files:
                if filemark in file:
                    matchedFlag = False
                    matchedInfo = list()

                    # 在文件中查找
                    src = os.path.join(source_root, file)

                    if not os.path.isfile(src):
                        continue

                    stream = open(src, 'r', encoding='utf-8', errors='ignore')
                    lines = stream.readlines()
                    for line in lines:
                        if wordmark in line:  # 匹配到wordmark
                            matchedFlag = True
                            matchedInfo.append(line)
                    stream.close()

                    # 打印信息
                    if matchedFlag:
                        print(Fore.RED + src)  # 打印wordmark所在文件路径（红色字体）
                        for line in matchedInfo:
                            print(Fore.RED + line)  # 打印wordmark所在行（红色字体）
                    else:
                        print(Fore.GREEN + src)  # 打印没有搜到wordmark的文件路径（绿色字体）

            for casefile in dirs:

                case_path = os.path.join(source_root, casefile)
                print(Fore.RESET + f'\n{case_path}')

                for root, dirs, files in os.walk(case_path):
                    for file in files:
                        if filemark in file:
                            # 初始化匹配标记
                            matchedFlag = False

                            # 初始化匹配行列表
                            matchedInfo = list()

                            # 在文件中查找
                            src = os.path.join(case_path, file)

                            if not os.path.isfile(src):
                                continue

                            stream = open(src, 'r', encoding='utf-8', errors='ignore')
                            for line in stream.readlines():
                                if wordmark in line:  # 匹配到wordmark
                                    matchedFlag = True
                                    matchedInfo.append(line)
                            stream.close()

                            # 打印信息
                            if matchedFlag:
                                print(Fore.RED + src)  # 打印wordmark所在文件路径（红色字体）
                                for line in matchedInfo:
                                    print(Fore.RED + line)  # 打印wordmark所在行（红色字体）
                            else:
                                print(Fore.GREEN + src)  # 打印没有搜到wordmark的文件路径（绿色字体）
                        else:
                            continue
    except Exception as e:
        logger.exception(e)


def test_pynmea2():
    src = r"D:\mfw\1_test\1.txt"

    stream = open(src, 'r', encoding='utf-8', errors='ignore')
    lines = stream.readlines()
    for line in lines:
        record = pynmea2.parse(line)
        print('1')
    stream.close()


def detect_restart():
    '''
    source_root > src_dir > file
    根据路径下.log文件，检测BK接收机是否发生重启动
    '''

    from colorama import Fore, Back, Style

    try:
        # source_root = input('info file root: ')
        source_root = r'\\192.168.60.217\Exchang_in\Common\WTR_LOG\23.10.07_166023395\data_spp_replay'
        # source_root = r'\\192.168.60.217\Exchang_in\Common\WTR_LOG\23.07.03_166023265\稳定性板子重启'
        # source_root = r'D:\mfw\1_test\tracking_nmea\2023717'

        for item0 in os.listdir(source_root):
            item_path = os.path.join(source_root, item0)

            if os.path.isdir(item_path):
                src_dir = item_path
                for item in os.listdir(src_dir):
                    item_path = os.path.join(src_dir, item)

                    if os.path.isfile(item_path) and item_path.endswith('log'):
                        src = item_path
                        stream = open(src, 'r', encoding='utf-8', errors='ignore')
                        lines = stream.readlines()
                        cnt_NormalVersionPrint = 0
                        error_flag = False
                        for line in lines:
                            if line.startswith('$POLRS'):
                                pos_status = line.split(',')[2][0]
                                if pos_status == '0' or pos_status == '2':  # 发命令或者正常定位
                                    cnt_NormalVersionPrint += 1

                                elif cnt_NormalVersionPrint >= 3 and pos_status == '1':  # 发生重启
                                    print(Fore.RED + src)
                                    print(Fore.RED + line)
                                    error_flag = True
                                elif cnt_NormalVersionPrint < 3 and pos_status == '1':  # 正常
                                    cnt_NormalVersionPrint = 0
                        stream.close()
                        # if error_flag is False:
                        #     print(Fore.GREEN + src)
                    else:
                        continue

                print('Checked: ' + src_dir)
            else:continue


    except Exception as e:
        logger.exception(e)


# 1.历元数检查
# xlsxfile = r'D:\mfw\练习\test_kml\case_FILE045_Dynamic\Compare\case_FILE045_Dynamic.xlsx'
# try:
#     check_epo(xlsxfile)
# except Exception as e:
#     logger.exception(e)

# 2.SPP回放检查
spp_check()

# 3.检查差分龄期
# check_rtk_delay()

# 4.剔除GIGGA
# delete_exgga()

# 5.统计cn0
# get_cn0()

# 6.修改nmea某项内容
# change_nmea_value('50000', 'GNGGA', 9)
# change_nmea_value('50000', 'GNRMC', )

# 7.批量检查fw_info.log中是否有关键字‘nan’
# search('nan', 'fw_info.log')
#
#
# 7.批量检查fw_info.log中是否有关键字‘nan’
# search('$CUST_9,255,255', '_fw_info')
# search('CMD_CTRL_RESET_WTD 1', '_fw_info')

# 8.
# test_pynmea2()

# 9.批量检查fw_info.log是否存在重启
# detect_restart()

# do_xl()
